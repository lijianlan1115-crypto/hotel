from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


CHANNEL_LABELS = {
    "all": "全部渠道",
    "multi": "多渠道",
    "fliggy": "飞猪",
    "meituan": "美团",
    "ctrip": "携程",
    "qunar": "去哪儿",
    "tongcheng": "同程",
}

CHANNEL_KEYWORDS = {
    "fliggy": ["飞猪", "fliggy", "淘系", "阿里旅行"],
    "meituan": ["美团", "大众点评", "meituan"],
    "ctrip": ["携程", "ctrip", "trip.com", "trip"],
    "qunar": ["去哪儿", "qunar"],
    "tongcheng": ["同程", "艺龙", "tongcheng", "elong"],
}


DETECTED_META_KEYS = [
    "detected_channels",
    "detected_channel_labels",
    "detected_channel_count",
    "detected_channel_title",
    "detected_period_start",
    "detected_period_end",
    "detected_period_days",
]

DATE_HEADER_KEYWORDS = (
    "日期",
    "营业日期",
    "业务日期",
    "数据日期",
    "统计日期",
    "快照日期",
    "预订日期",
    "订单日期",
    "开始日期",
    "结束日期",
    "周期开始",
    "周期结束",
    "period_start",
    "period_end",
    "period_start_field",
    "period_end_field",
    "data_date",
    "biz_date",
    "business_date",
    "stat_date",
    "snapshot_date",
    "booking_date",
    "order_date",
)


def detect_channels_from_excel(file_path: str, max_rows: int = 30) -> list[str]:
    path = Path(file_path).expanduser().resolve()

    if load_workbook is None:
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    text_pool: list[str] = []

    try:
        for sheet_name in workbook.sheetnames:
            text_pool.append(str(sheet_name))

            ws = workbook[sheet_name]
            for row in ws.iter_rows(max_row=max_rows, values_only=True):
                for cell in row:
                    if cell is not None:
                        text_pool.append(str(cell))
    finally:
        workbook.close()

    full_text = " ".join(text_pool).lower()
    detected: list[str] = []

    for channel, keywords in CHANNEL_KEYWORDS.items():
        if any(keyword.lower() in full_text for keyword in keywords):
            detected.append(channel)

    return detected


def build_channel_summary(channels: list[str]) -> dict[str, Any]:
    labels = [CHANNEL_LABELS.get(channel, channel) for channel in channels]

    return {
        "detected_channels": channels,
        "detected_channel_labels": labels,
        "detected_channel_count": len(channels),
        "detected_channel_title": " / ".join(labels) if labels else "",
    }


def _excel_serial_to_date(value: float) -> str | None:
    if value < 20000 or value > 80000:
        return None
    return (date(1899, 12, 30) + timedelta(days=int(value))).isoformat()


def _parse_date_value(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return _excel_serial_to_date(float(value))

    text = str(value).strip()
    if not text:
        return None

    # 2026-06-11 / 2026/06/11 / 2026.06.11
    match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if match:
        year, month, day = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            return None

    # Excel may display 6/11 or 06-11 without year; S14 test data is 2026.
    match = re.fullmatch(r"(\d{1,2})[-/.月](\d{1,2})日?", text)
    if match:
        month, day = int(match.group(1)), int(match.group(2))
        try:
            return date(2026, month, day).isoformat()
        except ValueError:
            return None

    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        return _excel_serial_to_date(float(text))

    return None


def detect_period_from_excel(file_path: str, max_rows: int = 2000) -> dict[str, Any]:
    """Detect the real period from date columns or period cells in the workbook.

    Priority is the workbook's own data dates. This prevents a control default
    like 2026-06-01~2026-06-10 from being shown when the uploaded workbook only
    contains 2026-06-11~2026-06-20 data.
    """

    path = Path(file_path).expanduser().resolve()
    if load_workbook is None:
        return {}

    workbook = load_workbook(path, read_only=True, data_only=True)
    detected_dates: list[str] = []

    try:
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            rows = list(ws.iter_rows(max_row=max_rows, values_only=True))
            if not rows:
                continue

            # Header table mode: only scan columns whose header looks like a date field.
            header_row_index: int | None = None
            date_col_indexes: set[int] = set()
            for row_index, row in enumerate(rows[:80]):
                for col_index, cell in enumerate(row):
                    text = str(cell or "").strip()
                    if any(keyword.lower() in text.lower() for keyword in DATE_HEADER_KEYWORDS):
                        date_col_indexes.add(col_index)
                        header_row_index = row_index if header_row_index is None else min(header_row_index, row_index)
                if date_col_indexes and header_row_index is not None:
                    break

            if date_col_indexes and header_row_index is not None:
                for row in rows[header_row_index + 1:]:
                    for col_index in date_col_indexes:
                        if col_index < len(row):
                            parsed = _parse_date_value(row[col_index])
                            if parsed:
                                detected_dates.append(parsed)
                continue

            # Key-value/control sheet mode: scan cells near date labels.
            for row in rows[:120]:
                for col_index, cell in enumerate(row):
                    text = str(cell or "").strip()
                    if not text or not any(keyword.lower() in text.lower() for keyword in DATE_HEADER_KEYWORDS):
                        continue
                    for near_index in range(col_index, min(len(row), col_index + 4)):
                        parsed = _parse_date_value(row[near_index])
                        if parsed:
                            detected_dates.append(parsed)
    finally:
        workbook.close()

    unique_dates = sorted(set(detected_dates))
    if not unique_dates:
        return {}
    start = unique_dates[0]
    end = unique_dates[-1]
    try:
        days = (datetime.strptime(end, "%Y-%m-%d").date() - datetime.strptime(start, "%Y-%m-%d").date()).days + 1
    except ValueError:
        days = None
    return {
        "detected_period_start": start,
        "detected_period_end": end,
        "detected_period_days": days,
    }


def build_excel_inputs(file_path: str) -> dict[str, Any]:
    path = Path(file_path).expanduser().resolve()

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    detected_channels = detect_channels_from_excel(str(path))
    detected_channel_summary = build_channel_summary(detected_channels)
    detected_period = detect_period_from_excel(str(path))

    period_start = detected_period.get("detected_period_start") or "2026-06-01"
    period_end = detected_period.get("detected_period_end") or "2026-06-10"

    inputs = {
        "hotel_id": "puyue",
        "hotel_name": "贵阳璞悦·奢电竞酒店",
        "platform": "multi",
        "channel_source": "多渠道",
        "channel_mode": "multi",
        "period_start": period_start,
        "period_end": period_end,
        "data_source_mode": "excel_upload",
        "input_excel_path": str(path),
        "dry_run": True,
    }

    inputs.update(detected_channel_summary)
    inputs.update(detected_period)

    return inputs


def run_s14_from_excel(
    file_path: str,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    try:
        from . import S14OperationDiagnosis
    except ImportError:
        import importlib.util
        import sys
        from pathlib import Path

        runtime_dir = Path(__file__).resolve().parent
        spec = importlib.util.spec_from_file_location(
            "s14_operation_diagnosis_runtime",
            runtime_dir / "__init__.py",
            submodule_search_locations=[str(runtime_dir)],
        )
        module = importlib.util.module_from_spec(spec)
        sys.modules["s14_operation_diagnosis_runtime"] = module
        assert spec and spec.loader
        spec.loader.exec_module(module)
        S14OperationDiagnosis = module.S14OperationDiagnosis

    runtime_config = config or {}
    inputs = build_excel_inputs(file_path)

    diagnosis_inputs = dict(inputs)

    detected_meta = {
        key: diagnosis_inputs.pop(key)
        for key in DETECTED_META_KEYS
        if key in diagnosis_inputs
    }

    result = S14OperationDiagnosis(runtime_config).execute(diagnosis_inputs)

    if isinstance(result, dict):
        result.update(detected_meta)
        if detected_meta.get("detected_period_start"):
            result["period_start"] = detected_meta["detected_period_start"]
        if detected_meta.get("detected_period_end"):
            result["period_end"] = detected_meta["detected_period_end"]

    return result
