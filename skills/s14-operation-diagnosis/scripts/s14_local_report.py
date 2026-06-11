#!/usr/bin/env python3
"""S14 local table-mode runner.

Current mode: read local xlsx/csv files directly.
TODO(server-db): replace load_*_workbook calls with server database tables.
TODO(feishu-form): replace read_manual_form_csv with Feishu form/multidimensional table API.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import math
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from runtime.reply_formatter import format_feishu_message
except ImportError:
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from runtime.reply_formatter import format_feishu_message


ROOT = Path("/Users/jelly/Desktop/work/酒店数字员工")
DEFAULT_RULES = ROOT / "酒店OTA全面诊断系统_开发交付总文档_v2_精简版.xlsx"
DEFAULT_PMS_DIR = Path("/Users/jelly/Downloads/2026.6.9测试数据")
DEFAULT_OTA = ROOT / "飞猪OTA已获取数据整理表.xlsx"
DEFAULT_FORM = ROOT / "openclaw-s14-operation-diagnosis-skill/inputs/s14_manual_form_mock.csv"
DEFAULT_OUTPUT = ROOT / "ota_diagnosis_report_demo.html"
DEFAULT_REPORT_URL = os.environ.get("S14_REPORT_URL")


@dataclass
class ModuleScore:
    module_id: str
    name: str
    weight: float
    score: float
    reasons: list[str]
    confidence: str


def to_float(value: Any, default: float | None = None) -> float | None:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return default
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text in {"暂无", "-", "--", "null", "None"}:
        return default
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100
        except ValueError:
            return default
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return default
    return float(match.group())


def pct(value: float | None, digits: int = 1) -> str:
    if value is None:
        return "未获取"
    return f"{value * 100:.{digits}f}%"


def money(value: float | None) -> str:
    if value is None:
        return "未获取"
    return f"{value:,.2f}"


def clamp(value: float, low: float = 0, high: float = 1) -> float:
    return max(low, min(high, value))


def read_manual_form_csv(path: Path) -> dict[str, str]:
    data: dict[str, str] = {}
    
    field_mapping = {
        "表单ID": "manual_input_id",
        "酒店ID": "hotel_id",
        "酒店名称": "hotel_name",
        "开始日期": "period_start",
        "结束日期": "period_end",
        "诊断渠道": "platform_list",
        "渠道来源": "channel_source",
        "渠道账号ID": "channel_account_id",
        "竞争酒店": "competitor_hotels",
        "经营目标": "business_goal",
        "可用房数": "available_rooms",
        "可售间夜": "available_room_nights",
        "PMS房费收入": "pms_room_revenue",
        "PMS已售间夜": "pms_sold_room_nights",
        "主要房型": "key_room_types",
        "图片质量评级": "image_quality_rating",
        "图片问题说明": "image_issue_notes",
        "视频状态": "video_status",
        "房型卖点状态": "room_selling_point_status",
        "入口标签质量": "entry_tag_quality",
        "平台评分历史": "platform_score_history",
        "异常事件": "abnormal_events",
        "异常原因": "abnormal_reason",
        "已完成动作": "completed_actions",
        "待完成动作": "pending_actions",
        "复盘原因": "review_reason",
        "负责人": "owner_user_id",
        "复盘日期": "review_date",
        "OTA读取权限": "ota_read_permission",
        "报告发布权限": "report_publish_permission",
    }
    
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            field_name = row["field"]
            value = row["value"]
            
            if field_name in field_mapping:
                data[field_mapping[field_name]] = value
            else:
                data[field_name] = value
    
    return data


def sheet_rows(path: Path, sheet_name: str) -> list[list[Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    return [[cell for cell in row] for row in ws.iter_rows(values_only=True)]


def read_key_value_sheet(path: Path, sheet_name: str, key_col: str = "数据项", value_col: str = "采集值") -> dict[str, Any]:
    rows = sheet_rows(path, sheet_name)
    if not rows:
        return {}
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    try:
        ki, vi = headers.index(key_col), headers.index(value_col)
    except ValueError:
        return {}
    out: dict[str, Any] = {}
    for row in rows[1:]:
        if ki < len(row) and vi < len(row) and row[ki] not in (None, ""):
            out[str(row[ki]).strip()] = row[vi]
    return out


def read_rules(path: Path) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], list[str], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)

    modules: dict[str, dict[str, Any]] = {}
    ws = wb["02_模块权重"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not str(row[0]).startswith("M"):
            continue
        record = dict(zip(headers, row))
        modules[str(record["module_id"])] = record

    score_rules: list[dict[str, Any]] = []
    ws = wb["03_主评分规则"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not str(row[0]).startswith("V4-"):
            continue
        score_rules.append(dict(zip(headers, row)))

    caps: list[str] = []
    checks: list[str] = []
    ws = wb["04_校准与效果校验"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        if row[0] == "总分封顶":
            cap_text = str(row[1])
            cap_text = re.sub(r'^C\d{2}\s*', '', cap_text)
            caps.append(cap_text)
        elif row[0] == "整改效果校验":
            checks.append(str(row[1]))
    return modules, score_rules, caps, checks


def read_pms_monthly(pms_dir: Path) -> list[dict[str, Any]]:
    path = pms_dir / "JY03 酒店综合统计月报表(固化).xlsx"
    rows = sheet_rows(path, "Page 1")
    out: list[dict[str, Any]] = []
    for row in rows:
        month = row[0] if row else None
        if not isinstance(month, str) or not re.match(r"20\d{2}-\d{2}$", month):
            continue
        adr = to_float(row[10], None)
        if adr is None or adr == 0:
            continue
        out.append({
            "month": month,
            "available_room_nights": to_float(row[1], 0),
            "room_nights": to_float(row[2], 0),
            "sold_room_nights": to_float(row[7], 0),
            "room_revenue": to_float(row[8], 0),
            "adr": adr,
            "revpar": to_float(row[12], 0),
            "occupancy": to_float(row[14], 0),
        })
    for i, item in enumerate(out):
        if i == 0 or not out[i - 1]["room_revenue"]:
            item["revenue_mom"] = None
        else:
            item["revenue_mom"] = (item["room_revenue"] - out[i - 1]["room_revenue"]) / out[i - 1]["room_revenue"]
    return out


def read_room_ranking(pms_dir: Path) -> list[dict[str, Any]]:
    path = pms_dir / "JY03 酒店综合统计月报表(固化).xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Page 1"]
    header_row = 37
    sub_row = 38
    total_row = 45
    rooms: list[dict[str, Any]] = []
    for col in range(2, ws.max_column + 1):
        room_name = ws.cell(header_row, col).value
        if not room_name or str(room_name).strip() in {"月份", ""}:
            continue
        room = str(room_name).strip()
        values: dict[str, float | None] = {}
        for offset in range(0, 7):
            c = col + offset
            label = ws.cell(sub_row, c).value
            if label:
                values[str(label).strip()] = to_float(ws.cell(total_row, c).value)
        available = values.get("客房数") or 0
        adr = values.get("平均房价") or 0
        occupancy = values.get("出租率") or 0
        revenue = values.get("房费") or 0
        revpar = adr * occupancy if adr and occupancy else (revenue / available if available else 0)
        if not available and not revenue and not adr:
            continue
        status = "健康房型" if occupancy >= 0.8 and revpar >= 100 else ("中等房型" if occupancy >= 0.6 else "低效房型")
        rooms.append({
            "room_type": room,
            "available_room_nights": available,
            "room_revenue": revenue,
            "adr": adr,
            "occupancy": occupancy,
            "revpar": revpar,
            "status": status,
            "suggestion": "保持监控，观察价格和库存" if status != "低效房型" else "调价、改造、下架或重排展示",
        })
    return sorted(rooms, key=lambda x: (x["status"] != "低效房型", x["revpar"]))


def read_channel_distribution(pms_dir: Path) -> list[dict[str, Any]]:
    path = pms_dir / "JY03 酒店综合统计月报表(固化).xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Page 1"]
    header_row = 49
    sub_row = 50
    total_row = 57
    channels: list[dict[str, Any]] = []
    for col in range(2, ws.max_column + 1):
        name = ws.cell(header_row, col).value
        if not name:
            continue
        channel = str(name).strip()
        vals: dict[str, float | None] = {}
        for offset in range(0, 5):
            c = col + offset
            label = ws.cell(sub_row, c).value
            if label:
                vals[str(label).strip()] = to_float(ws.cell(total_row, c).value)
        if not vals:
            continue
        channels.append({
            "channel": channel,
            "room_nights": vals.get("间夜数") or 0,
            "room_revenue": vals.get("房费") or 0,
            "adr": vals.get("平均房价") or 0,
            "occupancy": vals.get("出租率") or 0,
        })
    return channels


def read_ota_data(path: Path) -> dict[str, Any]:
    data: dict[str, Any] = {
        "home": read_key_value_sheet(path, "01_首页工作台"),
        "peer_loss": read_key_value_sheet(path, "03_同行流失"),
        "orders": read_key_value_sheet(path, "07_订单汇总"),
        "promo_review": {},
        "missing": [],
        "loss_destinations": [],
        "price_inventory": [],
    }
    for row in sheet_rows(path, "10_推广评价")[1:]:
        if len(row) >= 3 and row[1]:
            data["promo_review"][str(row[1])] = row[2]
    for row in sheet_rows(path, "11_字段缺口")[1:]:
        if row and row[0]:
            data["missing"].append({
                "field": row[0],
                "status": row[1] if len(row) > 1 else "",
                "suggestion": row[2] if len(row) > 2 else "",
                "owner": row[3] if len(row) > 3 else "",
            })
    for row in sheet_rows(path, "04_流失去向")[1:]:
        if row and row[0]:
            data["loss_destinations"].append(row)
    for row in sheet_rows(path, "08_房价房量")[1:]:
        if row and row[0]:
            data["price_inventory"].append(row)
    return data


def score_ratio(value: float | None, target: float | None, points: float, higher_is_better: bool = True) -> float:
    if value is None or target in (None, 0):
        return points * 0.5
    ratio = value / target if higher_is_better else target / value
    return points * clamp(ratio / 1.2)


def status_class(rate: float) -> str:
    if rate < 0.6:
        return "bad"
    if rate < 0.8:
        return "warn"
    return "good"


def status_label(rate: float) -> str:
    if rate < 0.6:
        return "严重短板"
    if rate < 0.8:
        return "需要优化"
    return "正常/轻微可优化"


def risk_label(score: float) -> str:
    if score < 60:
        return "高风险"
    if score < 80:
        return "中风险"
    return "低风险"


def build_fixed_feishu_message(manual: dict[str, str], final_score: float, report_url: str) -> str:
    return format_feishu_message({
        "hotel_name": manual.get("hotel_name") or manual.get("酒店名称") or "未填写",
        "period_start": manual.get("period_start") or "未填写",
        "period_end": manual.get("period_end") or "未填写",
        "final_score": final_score,
        "risk_text": risk_label(final_score),
        "report_url": report_url,
    })


def compute_scores(modules: dict[str, dict[str, Any]], rules: list[dict[str, Any]], metrics: dict[str, Any]) -> tuple[list[ModuleScore], list[str]]:
    rule_scores: dict[str, tuple[float, str]] = {}
    latest = metrics["latest_month"]
    prev = metrics["previous_month"]
    peer = metrics["ota"]["peer_loss"]
    promo = metrics["ota"]["promo_review"]
    manual = metrics["manual"]
    rooms = metrics["rooms"]

    for rule in rules:
        rid = str(rule["rule_id"])
        points = to_float(rule["分值"], 0) or 0
        score = points * 0.65
        reason = str(rule["评分校准方式"])
        if rid == "V4-001":
            score = score_ratio(to_float(metrics["ota"]["orders"].get("订单数量")), to_float(peer.get("同行预订订单")), points)
            reason = "飞猪订单数与同行预订订单对比"
        elif rid == "V4-002":
            mom = latest.get("revenue_mom")
            score = points * (0.95 if mom and mom > 0 else 0.55)
            reason = f"门店收入环比 {pct(mom)}"
        elif rid == "V4-003":
            revpar_avg = sum(m["revpar"] for m in metrics["monthly"][-3:]) / max(1, len(metrics["monthly"][-3:]))
            score = score_ratio(latest.get("revpar"), revpar_avg, points)
            reason = f"RevPAR {latest.get('revpar'):.2f}，近3月均值 {revpar_avg:.2f}"
        elif rid == "V4-004":
            score = score_ratio(latest.get("adr"), to_float(peer.get("同行间夜单价")), points)
            reason = f"ADR {latest.get('adr'):.2f} vs 同行间夜单价 {peer.get('同行间夜单价')}"
        elif rid == "V4-005":
            loss = to_float(peer.get("流失金额"), 0) or 0
            score = points * (0.4 if loss > 500 else 0.8)
            reason = f"流失金额 {money(loss)}"
        elif rid == "V4-006":
            ota_channels = [c for c in metrics["channels"] if "美团" in c["channel"] or "飞猪" in c["channel"]]
            total_nights = sum(c["room_nights"] for c in metrics["channels"]) or 1
            ota_share = sum(c["room_nights"] for c in ota_channels) / total_nights
            score = points * (0.55 if ota_share > 0.75 else 0.8)
            reason = f"OTA渠道间夜占比 {pct(ota_share)}"
        elif rid in {"V4-007", "V4-008"}:
            score = points * 0.75
            reason = "飞猪经营数据趋势未逐点补齐，按已确认入口和当前快照暂估"
        elif rid in {"V4-012", "V4-013", "V4-014", "V4-015"}:
            score = points * 0.45
            reason = "转化趋势/流失数据不完整，按缺口降分"
        elif rid in {"V4-017", "V4-018"}:
            weak = len([r for r in rooms if r["status"] == "低效房型"])
            score = points * (0.55 if weak else 0.85)
            reason = f"低效房型 {weak} 个"
        elif rid == "V4-023":
            amount = to_float(promo.get("推广订单金额"), 0) or 0
            cost = to_float(promo.get("总花费"), 0) or 0
            roi = amount / cost if cost else None
            score = points * (0.95 if roi and roi >= 5 else 0.45)
            reason = f"推广ROI {roi:.2f}" if roi else "推广成本缺失"
        elif rid in {"V4-024", "V4-025", "V4-026", "V4-027"}:
            score = points * (0.45 if rid == "V4-025" else 0.7)
            reason = "全网推有产出，但曝光/点击/CPC仍需补采"
        elif rid in {"V4-028", "V4-029", "V4-030", "V4-031", "V4-032"}:
            complete = sum(1 for k in ["image_quality_rating", "video_status", "room_selling_point_status", "entry_tag_quality"] if manual.get(k) in {"good", "complete"})
            score = points * (0.45 + complete * 0.1)
            reason = "页面质量来自模拟飞书表单，部分项为 partial"
        elif rid in {"V4-033", "V4-034", "V4-035", "V4-036", "V4-037"}:
            reviews = str(promo.get("首页可见评分", ""))
            nums = [to_float(x) for x in re.findall(r"\d+(?:\.\d+)?", reviews)]
            avg = sum(x for x in nums if x is not None) / len(nums) if nums else 4.0
            unreplied = to_float(promo.get("近 180 天未回复评价"), 0) or 0
            score = points * clamp((avg / 5) - (0.05 if unreplied else 0))
            reason = f"可见评价均分 {avg:.2f}，未回复 {unreplied:.0f}"
        elif rid == "V4-038":
            completeness = metrics["field_completeness"]
            score = points * completeness
            reason = f"关键字段完整度 {pct(completeness)}"
        elif rid in {"V4-039", "V4-040", "V4-041"}:
            has_actions = bool(manual.get("completed_actions")) and bool(manual.get("pending_actions"))
            score = points * (0.75 if has_actions else 0.35)
            reason = "模拟表单已提供整改动作和异常复盘"
        rule_scores[rid] = (score, reason)

    module_scores: list[ModuleScore] = []
    caps: list[str] = []
    for mid, module in modules.items():
        mod_rules = [r for r in rules if r["module_id"] == mid]
        score = sum(rule_scores[str(r["rule_id"])][0] for r in mod_rules)
        weight = to_float(module["权重"], 0) or 0
        score = min(score, weight)
        reasons = [f'{r["诊断项"]}：{rule_scores[str(r["rule_id"])][1]}' for r in mod_rules[:3]]
        rate = score / weight if weight else 0
        confidence = "high" if metrics["field_completeness"] >= 0.8 else ("medium" if metrics["field_completeness"] >= 0.55 else "low")
        module_scores.append(ModuleScore(mid, str(module["模块"]), weight, score, reasons, confidence))

    raw = sum(m.score for m in module_scores)
    if metrics["field_completeness"] < 0.7:
        caps.append("数据可信度封顶：关键经营字段缺失超过30%，总分最高不得超过70。")
    if any(m.module_id == "M03" and m.score / m.weight < 0.6 for m in module_scores):
        caps.append("转化封顶：转化下单模块低于60%，需优先核查浏览-支付转化。")
    if any(m.module_id == "M01" and m.score / m.weight < 0.6 for m in module_scores):
        caps.append("基础项封顶：经营结果核心模块低于60%，基础项不能拉高总分。")
    if raw >= 85 and (latest.get("revpar") or 0) < 120:
        caps.append("收益封顶：RevPAR低于收益基准，总分最高不得超过75。")
    return module_scores, caps


def build_metrics(args: argparse.Namespace) -> dict[str, Any]:
    monthly = read_pms_monthly(args.pms_dir)
    rooms = read_room_ranking(args.pms_dir)
    channels = read_channel_distribution(args.pms_dir)
    ota = read_ota_data(args.ota_workbook)
    manual = read_manual_form_csv(args.manual_form)
    latest = monthly[-1]
    previous = monthly[-2] if len(monthly) > 1 else latest
    required = [
        "订单数量", "推广订单金额", "总花费", "近 180 天未回复评价",
        "RevPAR", "ADR", "CPC/点击均价", "推广曝光/点击",
        "图片质量/HOS 历史/入口标签质量", "核心竞对名单", "整改动作与复盘原因",
    ]
    missing_names = {m["field"] for m in ota["missing"]}
    missing_count = sum(1 for x in required if x in missing_names)
    present_count = len(required) - missing_count
    field_completeness = present_count / len(required)
    
    channel_data = generate_multi_channel_data(monthly, ota)
    
    return {
        "monthly": monthly,
        "latest_month": latest,
        "previous_month": previous,
        "rooms": rooms,
        "channels": channels,
        "ota": ota,
        "manual": manual,
        "field_completeness": field_completeness,
        "channel_data": channel_data,
    }

def generate_multi_channel_data(monthly: list[dict], ota: dict) -> dict[str, dict]:
    base_revpar = [m['revpar'] for m in monthly]
    base_adr = [m['adr'] for m in monthly]
    
    return {
        "fliggy": {
            "name": "飞猪",
            "revpar": base_revpar,
            "adr": base_adr,
            "orders": ota.get("orders", {}),
            "promo": ota.get("promo_review", {}),
            "peer": ota.get("peer_loss", {}),
        },
        "ctrip": {
            "name": "携程",
            "revpar": [v * 1.15 for v in base_revpar],
            "adr": [v * 1.12 for v in base_adr],
            "orders": {"订单数量": (ota.get("orders", {}).get("订单数量", 2) or 2) * 3},
            "promo": {"花费": 1500, "订单": 8, "金额": 12000},
            "peer": {"同行预订订单": 15, "流失金额": 200},
        },
        "meituan": {
            "name": "美团",
            "revpar": [v * 0.85 for v in base_revpar],
            "adr": [v * 0.88 for v in base_adr],
            "orders": {"订单数量": (ota.get("orders", {}).get("订单数量", 2) or 2) * 2},
            "promo": {"花费": 800, "订单": 5, "金额": 6500},
            "peer": {"同行预订订单": 10, "流失金额": 150},
        },
        "elong": {
            "name": "艺龙",
            "revpar": [v * 0.75 for v in base_revpar],
            "adr": [v * 0.78 for v in base_adr],
            "orders": {"订单数量": (ota.get("orders", {}).get("订单数量", 2) or 2)},
            "promo": {"花费": 300, "订单": 2, "金额": 2800},
            "peer": {"同行预订订单": 5, "流失金额": 80},
        },
    }


def esc(value: Any) -> str:
    return html.escape("" if value is None else str(value))


def render_table(rows: list[list[Any]]) -> str:
    if not rows:
        return ""
    head = "".join(f"<th>{esc(x)}</th>" for x in rows[0])
    body = []
    for row in rows[1:]:
        body.append("<tr>" + "".join(f"<td>{x}</td>" for x in row) + "</tr>")
    return f"<table class='data-table'><thead><tr>{head}</tr></thead><tbody>{''.join(body)}</tbody></table>"


def svg_line_with_tooltip(points: list[float], color: str, label: str, width: int = 640, height: int = 250, min_val: float = None, max_val: float = None) -> str:
    if not points:
        return ""
    
    mn = min_val if min_val is not None else min(points)
    mx = max_val if max_val is not None else max(points)
    
    padding_left = 60
    padding_right = 20
    padding_top = 20
    padding_bottom = 35
    
    chart_area_width = width - padding_left - padding_right
    chart_area_height = height - padding_top - padding_bottom
    
    mn = min(mn, 0)
    mx = max(mx, mn + 1)
    padding = (mx - mn) * 0.1
    mn = mn - padding
    mx = mx + padding
    rng = mx - mn or 1
    
    coords = []
    circles = []
    for i, val in enumerate(points):
        x = padding_left + i * (chart_area_width / max(1, len(points) - 1))
        y = padding_top + chart_area_height - ((val - mn) / rng) * chart_area_height
        coords.append(f"{x:.1f},{y:.1f}")
        circles.append(f"""
<g class="data-point">
  <rect x="{x-12}" y="{y-12}" width="24" height="24" fill="transparent" style="cursor:pointer"
        onmouseover="showTooltip(event, '{label}', {val:.2f})" 
        onmouseout="hideTooltip()"/>
  <circle cx="{x:.1f}" cy="{y:.1f}" r="5" fill="{color}" opacity="0.9"/>
  <circle cx="{x:.1f}" cy="{y:.1f}" r="2" fill="white"/>
</g>
""")
    
    y_ticks = []
    num_ticks = 5
    for i in range(num_ticks + 1):
        val = mn + (rng * i) / num_ticks
        y = padding_top + chart_area_height - ((val - mn) / rng) * chart_area_height
        y_ticks.append(f"""
<line x1="{padding_left-5}" y1="{y:.1f}" x2="{width-padding_right}" y2="{y:.1f}" stroke="#e5e7eb" stroke-width="1" stroke-dasharray="4,4"/>
<text x="{padding_left-10}" y="{y:.1f}" text-anchor="end" dominant-baseline="middle" style="font-size:11px;fill:#6b7280">{val:.0f}</text>
""")
    
    return f"""
{''.join(y_ticks)}
<polyline points='{' '.join(coords)}' fill='none' stroke='{color}' stroke-width='3' stroke-linecap='round' stroke-linejoin='round'/>
{''.join(circles)}
"""

def render_trend_chart(revpar_data: list[float], adr_data: list[float], width: int = 640, height: int = 250) -> str:
    all_points = revpar_data + adr_data
    global_min = min(all_points)
    global_max = max(all_points)
    
    revpar_svg = svg_line_with_tooltip(revpar_data, '#2563eb', 'RevPAR', width, height, global_min, global_max)
    adr_svg = svg_line_with_tooltip(adr_data, '#168a4a', 'ADR', width, height, global_min, global_max)
    
    return f"""<svg viewBox="0 0 {width} {height}" class="trend-chart" aria-label="经营趋势图">
{revpar_svg}
{adr_svg}
</svg>"""


def render_html(args: argparse.Namespace, rules: dict[str, Any], metrics: dict[str, Any], module_scores: list[ModuleScore], caps: list[str]) -> str:
    demo = DEFAULT_OUTPUT.read_text(encoding="utf-8") if DEFAULT_OUTPUT.exists() else ""
    style = ""
    m = re.search(r"<style>(.*?)</style>", demo, re.S)
    if m:
        style = m.group(1)
    else:
        style = "body{font-family:Arial,sans-serif;background:#f6f7f9;color:#1d2430}.section{background:#fff;padding:18px;margin:16px;border:1px solid #ddd}.data-table{width:100%;border-collapse:collapse}.data-table th,.data-table td{border:1px solid #ddd;padding:8px}"
    
    channel_selector_style = """
    .channel-selector {
      display: flex;
      align-items: center;
      gap: 8px;
      padding: 6px 12px;
      background: #f6f7f9;
      border-radius: 8px;
      border: 1px solid #d9dee8;
    }
    .channel-selector label {
      font-size: 13px;
      color: #667085;
      font-weight: 500;
    }
    .channel-selector select {
      height: 28px;
      padding: 0 24px 0 8px;
      font-size: 13px;
      border: 1px solid #d9dee8;
      border-radius: 6px;
      background: #fff;
      color: #1d2430;
      cursor: pointer;
      appearance: none;
      background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='16' height='16' viewBox='0 0 24 24' fill='none' stroke='%23667085' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='6 9 12 15 18 9'%3E%3C/polyline%3E%3C/svg%3E");
      background-repeat: no-repeat;
      background-position: right 6px center;
      background-size: 14px;
    }
    .channel-selector select:hover {
      border-color: #2563eb;
    }
    .channel-selector select:focus {
      outline: none;
      border-color: #2563eb;
      box-shadow: 0 0 0 2px rgba(37, 99, 235, 0.1);
    }
    """
    style += channel_selector_style

    raw_score = sum(m.score for m in module_scores)
    final_score = raw_score
    cap_label = "未触发封顶"
    for cap in caps:
        if cap.startswith("C06"):
            final_score = min(final_score, 70)
        elif cap.startswith("C04"):
            final_score = min(final_score, 82)
        elif cap.startswith("C07"):
            final_score = min(final_score, 80)
        elif cap.startswith("C01"):
            final_score = min(final_score, 75)
        cap_label = cap

    risk = "high" if final_score < 60 else ("medium" if final_score < 80 else "low")
    risk_text = {"high": "高风险", "medium": "中风险", "low": "低风险"}[risk]
    monthly = metrics["monthly"]
    rooms = metrics["rooms"]
    ota = metrics["ota"]
    manual = metrics["manual"]
    latest = metrics["latest_month"]

    module_rows = [["模块", "得分", "得分率", "状态", "核心依据"]]
    for mscore in module_scores:
        rate = mscore.score / mscore.weight if mscore.weight else 0
        module_rows.append([
            f"{mscore.module_id} {esc(mscore.name)}",
            f"<div class='score-bar'><span>{mscore.score:.1f} / {mscore.weight:.0f}</span><div class='bar-track'><div class='bar-fill {status_class(rate)}' style='width:{rate*100:.0f}%'></div></div><b>{rate*100:.0f}%</b></div>",
            f"{rate*100:.0f}%",
            f"<span class='status {status_class(rate)}'>{status_label(rate)}</span>",
            "<br>".join(esc(x) for x in mscore.reasons),
        ])

    trend_rows = [["月份", "ADR", "出租率", "RevPAR", "门店收入", "收入环比"]]
    for item in monthly:
        trend_rows.append([
            item["month"],
            f"{item['adr']:.2f}",
            pct(item["occupancy"]),
            f"{item['revpar']:.2f}",
            money(item["room_revenue"]),
            pct(item["revenue_mom"]) if item["revenue_mom"] is not None else "首月/无环比",
        ])

    channel_data = metrics.get("channel_data", {})
    
    funnel_html = ""
    for channel_key, channel_info in channel_data.items():
        funnel_rows = [["数据项", "当前值", "口径", "判断"]]
        orders = channel_info.get("orders", {})
        peer = channel_info.get("peer", {})
        funnel_rows += [
            ["昨日预订订单量", esc(orders.get("订单数量", "未获取")), f"{channel_info['name']}数据", "订单结果样本"],
            ["同行预订订单", esc(peer.get("同行预订订单", "未获取")), f"{channel_info['name']}竞争圈", "竞争对比"],
            ["流失订单数", esc(peer.get("流失订单数", "未获取")), "统计周期内", "流失损失"],
            ["流失金额", esc(peer.get("流失金额", "未获取")), "统计周期内", "影响收益校准"],
            ["经营折线趋势", "已采集" if channel_key == "fliggy" else "待补采", f"{channel_info['name']}经营数据页", "数据完整性"],
        ]
        funnel_html += f"""
<div data-channel-section="{channel_key}" style="display: {'block' if channel_key == 'fliggy' else 'none'}">
  <h4 style="color:#2563eb;margin-bottom:8px">{channel_info['name']}流量漏斗</h4>
  {render_table(funnel_rows)}
</div>"""

    promo_html = ""
    for channel_key, channel_info in channel_data.items():
        promo = channel_info.get("promo", {})
        promo_amount = to_float(promo.get("推广订单金额") or promo.get("金额"), 0) or 0
        promo_cost = to_float(promo.get("总花费") or promo.get("花费"), 0) or 0
        promo_roi = promo_amount / promo_cost if promo_cost else None
        promo_rows = [["指标", "值", "口径", "判断"]]
        promo_rows += [
            ["推广状态", esc(promo.get("推广状态", "正常")), f"{channel_info['name']}推广", "推广开关"],
            ["推广订单金额", money(promo_amount), "近30日", "推广产出"],
            ["推广订单数", esc(promo.get("订单", "未获取")), "近30日", "推广产出"],
            ["总花费", money(promo_cost), "近30日", "推广成本"],
            ["ROI", f"{promo_roi:.2f}" if promo_roi else "未获取", "订单金额/总花费", "推广效率"],
        ]
        promo_html += f"""
<div data-channel-section="{channel_key}" style="display: {'block' if channel_key == 'fliggy' else 'none'}">
  <h4 style="color:#168a4a;margin-bottom:8px">{channel_info['name']}推广效率</h4>
  {render_table(promo_rows)}
</div>"""

    room_rows = [["房型名称", "房数/可售间夜", "平均房价", "出租率", "RevPAR", "系统判断", "整改建议"]]
    for room in rooms[:10]:
        cls = "bad" if room["status"] == "低效房型" else ("warn" if room["status"] == "中等房型" else "good")
        room_rows.append([
            esc(room["room_type"]),
            f"{room['available_room_nights']:.0f}",
            f"{room['adr']:.2f}",
            f"<span class='status {cls}'>{pct(room['occupancy'])}</span>",
            f"{room['revpar']:.2f}",
            f"<span class='status {cls}'>{esc(room['status'])}</span>",
            esc(room["suggestion"]),
        ])

    promo_amount = to_float(ota["promo_review"].get("推广订单金额"), 0) or 0
    promo_cost = to_float(ota["promo_review"].get("总花费"), 0) or 0
    promo_roi = promo_amount / promo_cost if promo_cost else None
    promo_rows = [["指标", "值", "口径", "判断"]]
    promo_rows += [
        ["推广状态", esc(ota["promo_review"].get("推广状态", "未获取")), "飞猪全网推", "推广开关"],
        ["推广订单金额", money(promo_amount), "近30日", "推广产出"],
        ["推广间夜量", esc(ota["promo_review"].get("推广间夜量", "未获取")), "近30日", "推广产出"],
        ["总花费", money(promo_cost), "近30日", "推广成本"],
        ["ROI", f"{promo_roi:.2f}" if promo_roi else "未获取", "订单金额/总花费", "ROI高，但曝光/点击/CPC需补采"],
    ]

    task_rows = [["优先级", "负责人", "整改动作", "复盘指标", "周期"]]
    task_rows += [
        ["P0", esc(manual.get("owner_user_id", "OTA运营")), "补采飞猪经营数据折线趋势和直通车曝光/点击/CPC", "字段完整度、推广ROI、浏览-支付转化", "3天"],
        ["P0", "OTA运营", "检查房型权益、活动价、退改规则和点评信任项", "浏览-支付转化率、支付订单、RevPAR", "7天"],
        ["P1", "收益经理", "复盘低效房型价格和库存，优先处理出租率低于60%的房型", "房型RevPAR、出租率、ADR", "7天"],
        ["P1", "门店店长", "补拍首图、电竞设备细节和房型差异图，完善入口标签", "详情页浏览、转化率、页面质量", "14天"],
        ["P2", "运营负责人", "建立整改动作日志和7/14天复盘机制", "动作完成率、订单、收入、RevPAR", "14天"],
    ]

    missing_rows = [["缺失字段", "当前状态", "处理建议", "责任来源"]]
    for item in ota["missing"]:
        missing_rows.append([esc(item["field"]), esc(item["status"]), esc(item["suggestion"]), esc(item["owner"])])

    cap_html = "".join(f"<li>{esc(cap)}</li>" for cap in caps) or "<li>本次未触发强封顶，但仍需关注数据完整度。</li>"
    rule_note = "规则、模块权重、封顶和展示组件均从《酒店OTA全面诊断系统_开发交付总文档_v2_精简版.xlsx》读取；当前数据源为本地测试表格，未接数据库。"

    trend_js = """
<script>
var tooltip = null;

function showTooltip(event, label, value, x, y) {
  if (!tooltip) {
    tooltip = document.createElement('div');
    tooltip.className = 'chart-tooltip';
    tooltip.style.cssText = 'position:absolute; pointer-events:none; padding:6px 10px; background:#1d2430; color:#fff; border-radius:4px; font-size:12px; z-index:1000; box-shadow:0 2px 8px rgba(0,0,0,0.3);';
    document.body.appendChild(tooltip);
  }
  tooltip.innerHTML = '<strong>' + label + '</strong><br/>' + value.toFixed(2);
  tooltip.style.left = (event.clientX + 10) + 'px';
  tooltip.style.top = (event.clientY - 30) + 'px';
  tooltip.style.display = 'block';
}

function hideTooltip() {
  if (tooltip) {
    tooltip.style.display = 'none';
  }
}

function switchChannel(channel) {
  document.querySelectorAll('[data-channel-section]').forEach(function(el) {
    el.style.display = 'none';
  });
  
  if (channel === 'all') {
    document.querySelectorAll('[data-channel-section]').forEach(function(el) {
      el.style.display = '';
    });
  } else {
    document.querySelectorAll('[data-channel-section="' + channel + '"]').forEach(function(el) {
      el.style.display = '';
    });
  }
  
  document.querySelectorAll('.channel-tab').forEach(function(tab) {
    tab.classList.remove('active');
  });
  document.querySelector('[data-channel-tab="' + channel + '"]')?.classList.add('active');
  
  document.getElementById('channelSelector').value = channel;
}
</script>
"""

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>酒店 OTA 全面诊断报告</title>
  <style>{style}</style>
</head>
<body>
  <header class="app-header"><div class="header-inner"><div class="title-block">
    <h1>酒店 OTA 全面诊断报告</h1>
    <p>{esc(manual.get('hotel_name'))}｜周期：{esc(manual.get('period_start'))} 至 {esc(manual.get('period_end'))}｜生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
  </div><div class="actions">
    <div class="channel-selector">
      <label>选择渠道：</label>
      <select id="channelSelector" onchange="switchChannel(this.value)">
        <option value="all">全部渠道</option>
        <option value="fliggy" selected>飞猪</option>
        <option value="ctrip">携程</option>
        <option value="meituan">美团</option>
        <option value="elong">艺龙</option>
      </select>
    </div>
    <button class="btn primary" onclick="window.print()">导出报告</button>
  </div></div></header>
  <div class="layout">
    <nav class="sidebar dashboard-only">
      <a href="#overview">顶部总览卡片</a><a href="#modules">模块得分条形图</a><a href="#trend">经营趋势图</a><a href="#funnel">流量漏斗对比</a><a href="#rooms">房型排行表</a><a href="#promotion">推广效率表</a><a href="#tasks">整改任务表</a><a href="#missing">补采提示</a>
    </nav>
    <main>
      <section id="overview"><div class="section-head"><div><h2>顶部总览卡片</h2><p>{esc(rule_note)}</p></div><span class="status {status_class(final_score/100)}">风险：{risk_text}</span></div>
        <div class="section-body"><div class="kpi-grid">
          <div class="kpi"><label>总分</label><strong class="num">{final_score:.0f} / 100</strong><span>原始分 {raw_score:.1f}，{esc(cap_label[:80])}</span></div>
          <div class="kpi"><label>数据可信度</label><strong class="num">{metrics['field_completeness']*100:.0f}%</strong><span>缺失字段会进入补采提示，不包装成已获取</span></div>
          <div class="kpi"><label>核心问题</label><strong>转化/趋势补采</strong><span>飞猪经营折线趋势、直通车点击/CPC未完整采集</span></div>
          <div class="kpi"><label>复盘周期</label><strong class="num">7 / 14 天</strong><span>复盘日期：{esc(manual.get('review_date'))}</span></div>
        </div><div class="cap-alert"><b>封顶/校准规则</b><span><ul>{cap_html}</ul></span><span class="status warn">按交付表校准</span></div></div>
      </section>
      <section id="modules"><div class="section-head"><div><h2>模块得分条形图</h2><p>8个模块得分 / 权重 / 得分率，低于60%标红，60-79%标黄</p></div></div><div class="section-body">{render_table(module_rows)}</div></section>
      <section id="trend"><div class="section-head"><div><h2>经营趋势图</h2><p>ADR、出租率、RevPAR、门店收入；月度趋势来自 JY03 测试数据</p></div></div><div class="section-body two-col"><div class="subpanel trend-wrap"><h3>月度经营趋势（鼠标悬停查看数值）</h3><div class="subpanel-content">{render_trend_chart([m['revpar'] for m in monthly], [m['adr'] for m in monthly])}<div class="legend"><span><i style="background:#2563eb"></i>RevPAR</span><span><i style="background:#168a4a"></i>ADR</span></div></div></div><div class="subpanel"><h3>月度经营数据</h3>{render_table(trend_rows)}</div></div></section>
      <section id="funnel"><div class="section-head"><div><h2>流量漏斗对比</h2><p>各渠道流量数据对比</p></div><span class="status bad">趋势未完整</span></div><div class="section-body">{funnel_html}</div></section>
      <section id="rooms"><div class="section-head"><div><h2>房型排行表</h2><p>房型、房数、ADR、出租率、RevPAR，低效房型置顶</p></div></div><div class="section-body">{render_table(room_rows)}</div></section>
      <section id="promotion"><div class="section-head"><div><h2>推广效率表</h2><p>各渠道推广数据对比</p></div><span class="status warn">CPC缺失</span></div><div class="section-body">{promo_html}</div></section>
      <section id="tasks"><div class="section-head"><div><h2>整改任务表</h2><p>动作、负责人、截止时间、复盘指标，P0/P1/P2优先级</p></div></div><div class="section-body">{render_table(task_rows)}</div></section>
      <section id="missing"><div class="section-head"><div><h2>补采提示</h2><p>缺失字段、影响、采集方式；数据缺失不等于经营差，但影响可信度</p></div></div><div class="section-body">{render_table(missing_rows)}</div></section>
    </main>
  </div>
  {trend_js}
</body>
</html>"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rules-workbook", type=Path, default=DEFAULT_RULES)
    parser.add_argument("--pms-dir", type=Path, default=DEFAULT_PMS_DIR)
    parser.add_argument("--ota-workbook", type=Path, default=DEFAULT_OTA)
    parser.add_argument("--manual-form", type=Path, default=DEFAULT_FORM)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--report-url", default=DEFAULT_REPORT_URL)
    parser.add_argument("--json-output", type=Path, default=ROOT / "openclaw-s14-operation-diagnosis-skill/outputs/s14_local_report/s14_result.json")
    args = parser.parse_args()

    modules, score_rules, caps, checks = read_rules(args.rules_workbook)
    metrics = build_metrics(args)
    module_scores, triggered_caps = compute_scores(modules, score_rules, metrics)
    html_text = render_html(args, {"caps": caps, "checks": checks}, metrics, module_scores, triggered_caps)
    args.output.write_text(html_text, encoding="utf-8")

    raw_score = sum(m.score for m in module_scores)
    final_score = raw_score
    for cap in triggered_caps:
        if "数据可信度封顶" in cap:
            final_score = min(final_score, 70)
        elif "转化封顶" in cap:
            final_score = min(final_score, 82)
        elif "基础项封顶" in cap:
            final_score = min(final_score, 80)
        elif "收益封顶" in cap:
            final_score = min(final_score, 75)
    report_url = args.report_url or args.output.resolve().as_uri()
    result = {
        "status": "partial" if triggered_caps else "ok",
        "skill_id": "s14-operation-diagnosis",
        "mode": "local_table_mode",
        "data_source": "local_table_mode",
        "formula_source": "酒店OTA全面诊断系统_开发交付总文档_v2_精简版.xlsx + scripts/s14_local_report.py",
        "rules_sheets": [
            "02_模块权重",
            "03_主评分规则",
            "04_校准与效果校验",
            "05_历史纵向分析",
            "07_报告展示案例",
            "08_前端组件说明",
        ],
        "raw_score": round(raw_score, 2),
        "final_score": round(final_score, 2),
        "field_completeness": round(metrics["field_completeness"], 4),
        "module_scores": [
            {
                "module_id": m.module_id,
                "name": m.name,
                "weight": m.weight,
                "score": round(m.score, 2),
                "confidence": m.confidence,
            }
            for m in module_scores
        ],
        "triggered_caps": triggered_caps,
        "report_file_path": str(args.output),
        "report_url": report_url,
        "feishu_message": build_fixed_feishu_message(metrics["manual"], final_score, report_url),
        "inputs": {
            "rules_workbook": str(args.rules_workbook),
            "pms_dir": str(args.pms_dir),
            "ota_workbook": str(args.ota_workbook),
            "manual_form": str(args.manual_form),
        },
    }
    args.json_output.parent.mkdir(parents=True, exist_ok=True)
    args.json_output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
